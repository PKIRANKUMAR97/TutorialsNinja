import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(filename,sheetname):
    wb = openpyxl.load_workbook(filename)
    sh = wb[sheetname]
    rows= sh.max_row
    return rows

def getColumnCount(filename,sheetname):
    wb = openpyxl.load_workbook(filename)
    sh=wb[sheetname]
    cols = sh.max_column
    return cols

def readData(filename,sheetname,row,col):
    wb=openpyxl.load_workbook(filename)
    sh=wb[sheetname]
    return sh.cell(row,col).value

def writeData(filename,sheetname,row,col,data_to_be_written):
    wb=openpyxl.load_workbook(filename)
    sh=wb[sheetname]
    sh.cell(row,col).value=data_to_be_written
    wb.save(filename)

